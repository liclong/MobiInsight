#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
Created on July 22 2020 @ HongKong
Author: Changlong Li
Email: liclong@mail.ustc.edu.cn
"""

import os
import time
import queue
import threading
import requests
import openpyxl

from time import sleep
from tqdm import tqdm

path = 'apk'
download_file = 'crawler.xlsx'
download_max = 0

header = {'User-Agent':
              'Mozilla/5.0 (Windows NT 6.1 WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.93 Safari/537.36'}


def inx(path):
    apk_list = []
    for info in (os.walk(path, topdown=True)):
        if info[2]:  # File exists in this directory
            for apk_check in info[2]:
                if '.apk' in apk_check:
                    apk_list.append(apk_check)
    return apk_list


def apk_install():
    count = 0
    for package_name in package_list:
        cmd = 'adb install ' + path + '/' + package_name
        print(cmd)
        if os.system(cmd) == 0:
            count = count + 1
        else:  # Warning: an apk is not installed successfully
            pass
    time.sleep(1)
    print("There are %d apks installed successfully in total" % (count))


def download_single_apk(apk_url_str):
    apk_name, apk_url = apk_url_str.split(";")
    # print(apk_url)
    save_path = os.path.join(curdir, path, "%s.apk" % apk_name)
    if not os.path.exists(save_path):  # avoid repeated downloading
        # print("Downloading  %s" % save_path)
        try:
            r = requests.get(apk_url, headers=header, allow_redirects=True, timeout=720)
            status_code = r.status_code
            if status_code == 200 or status_code == 206:
                with open(save_path, "wb") as hf:
                    hf.write(r.content)
        except:
            print("Error, can not download %s.apk" % apk_name)
            pass
    else:
        print("%s downloaded already!" % save_path)


class DownloadThread(threading.Thread):  # apk download in batch
    def __init__(self, q_job):
        self._q_job = q_job
        threading.Thread.__init__(self)
        self.result = 1

    def run(self):
        while True:
            if self._q_job.qsize() > 0:
                download_single_apk(self._q_job.get())
                self.result = 0
            else:
                break


def apk_download():
    q = queue.Queue(0)  # queue initialization

    excel = openpyxl.load_workbook(download_file)
    table = excel.active
    rows = table.max_row
    # download_max = rows
    for r in range(1, rows + 1):  # read excel from the second row, the second column
        apk_name = table.cell(row=r, column=1).value  # get apk_name
        apk_url = table.cell(row=r, column=2).value  # get apk_url
        print(apk_name + "; " + apk_url)
        temp_str = apk_name + ";" + apk_url
        q.put(temp_str)

    for i in range(10):  # create ten threads for downloading
        DownloadThread(q).start()


if __name__ == '__main__':

    curdir = os.getcwd()  # get current work directory

    if not os.path.exists(path):
        os.system("mkdir " + path)

    print('Package indexing')
    package_list = inx(path)
    if not package_list:
        print('No package is found, try to download online ...')
        excel = openpyxl.load_workbook(download_file)
        table = excel.active
        download_max = table.max_row
        apk_download()
        download_current = 0
        k_chance = 24
        print('prepare ...')
        while True:
            time.sleep(10)
            package_list = inx(path)
            if package_list:
                tmp_len = len(package_list)
            else:
                tmp_len = 0
            if tmp_len == download_current:
                k_chance = k_chance - 1
                if k_chance == 0:
                    print('Download time out ...')
                    break
            else:
                k_chance = 24
                download_current = tmp_len
                for i in tqdm(range(download_current)):
                    sleep(0.5)
                if download_current == download_max:
                    print('All of the %d apks are installed successfully.' % (download_current))
                    break
    print('There are %d apks found in dir: %s/' % (len(package_list), path))
    apk_install()
